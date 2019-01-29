# readCensusExcel.py - Tabulates population and number of census tracts for each county
#! python3

import openpyxl, pprint
print('Opening workbook...')

# using short or long file path depends on in which folder application (VSC) is opened, when exact folder with file is opened('parent') then only file name
wb = openpyxl.load_workbook('censuspopdata.xlsx')
#wb = openpyxl.load_workbook('C:\\Users\\14000322\\Documents\\Pliki Pythona\\Licho-repository\\inne_projekty\\automate_boring_stuff\\censuspopdata.xlsx')
#wb.get_sheet_names() #will show only 1 sheet in Workbook object and its name(used below)

sheet = wb['Population by Census Tract'] #getting spreadsheet from Workbook object

countyData = {}

# TODO: filling in countyData with each county's population and tracts
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract(means population counting)
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # making sure the key for this state exists
    countyData.setdefault(state, {}) #setdefault will do nothing if key exists
    # making sure the key for this county in this state exists
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # incrementing 'tracts' key and increasing the value of 'pop' key with each next row
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop) # 'pop' value is integer so..?
    

# TODO: opening a new text file and writing the contents of countyData to it
print('Writing results...')
# outputting data to a text file with pyton extension benefits with access to
# that data like a python module (import..) - helpful in interactive shell
resultFile = open('census2010.py', 'w')
# pprint.pformat is formatting a given data to pretty-printed representation
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')
