#! python
# excel2csv.py - converts all excel files in given directory to similarly
# named csv files.

import os
import csv
import openpyxl

for excelFile in os.listdir(os.getcwd()+'\\Excels'):
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(os.path.join(os.getcwd()+'\\Excels',excelFile))
        
        for sheetName in wb.get_sheet_names():
            # Loop through every sheet in the workbook.
            sheet = wb.get_sheet_by_name(sheetName)
            # Create the CSV filename from the Excel filename and sheet title.
            f_name = excelFile[:-5]
            csv_file = open(f_name + '_' + sheetName + '.csv', 'w', newline='')
            # Create the csv.writer object for this CSV file.
            csv_writer = csv.writer(csv_file)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                    
                # Write the rowData list to the CSV file.
                for row in rowData:
                    csv_writer.writerow(row)                   

            csv_file.close()
