# multiplication_table.py - program generating spreadsheet after calling from
# command line file name with integer parameter
'''Steps:
1. importing modules - needs to think which ones will be useful
2. assigning the parameter to use in command line - size of table 
3. activating spreadsheet
4. preparing the outer cells as edges to multiplying table insid
5. fulfilling the inside of the table with correct formulas
6. saving workbook and opening it

'''
import sys, os
import openpyxl as pxl
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter

num = int(sys.argv[1])
cells = num + 1

wb =pxl.Workbook()
sheet = wb.active
sheet.title = 'Multiplying Table'

width = 6
heigth = 32
# preparing format for 1 column and 1 row it multiplying tabel
make_bold = Font(bold=True, size=20)
make_12 = Font(size=12)
sheet.row_dimensions[1].heigth = heigth
sheet.column_dimensions['A'].width = width
while cells > 1:
    # write outer cells values and make them bold
    sheet.cell(row=cells, column=1).value = cells - 1
    sheet.cell(row=cells, column=1).font = make_bold
    #sheet.row_dimensions[cells].heigth = heigth
    sheet.cell(row=1, column=cells).value = cells -1
    sheet.cell(row=1, column=cells).font = make_bold
    sheet.column_dimensions[get_column_letter(cells)].width = width
    cells -= 1

# populating the inner cells with correct formula
for row_ind in range(1, num+1):
    for column_ind in range(1, num+1):
        sheet.cell(row=row_ind+1, column=column_ind+1).value = row_ind * column_ind
        sheet.cell(row=row_ind+1, column=column_ind+1).font = make_12

wb.save('multi_table.xlsx')

# opening spreadsheet in default application
os.startfile('multi_table.xlsx')
        
