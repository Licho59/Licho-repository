# blank_row_inserter.py - file operated from command line

"""
Replicate a spreadsheet in CWD but with N blank rows inserted at chosen place.

Usage: blank_row_inserter.py takes 3 arguments
    <Name> - Name of the spreadsheet to add blank rows to
        <Start> - The row to start inserting the blank lines
            <Number> - The number of blank lines to insert

e.g. python blank_row_inserter.py test.xlsx 3 2 - would insert 2 blank rows
starting at row 3 into a copy of 'test.xlsx'
"""
import sys
import openpyxl
'''
name='produceSales.xlsx'
blank_start = 2
blank_length = 4

'''
name = sys.argv[1]
blank_start = int(sys.argv[2])
blank_length = int(sys.argv[3])

# reading the spreadsheet into a list of row data lists with cell's
# values of the spreadsheet
print('Reading spreadsheet data...')

wb = openpyxl.load_workbook(name)
sheet = wb.active
max_col = sheet.max_column
max_row = sheet.max_row

rows = []
for row in range(1, max_row+1):
    data = []
    for cell in range(1, max_col+1):
        data.append(sheet.cell(row=row, column=cell).value)
    rows.append(data)

# writing rows into the new spreadsheet with blank lines inserted
print('Inserting blank rows...')

wb = openpyxl.Workbook()
sheet = wb.active

for row in range(1, blank_start):
    for cell in range(1, max_col+1):
        sheet.cell(row=row, column=cell).value = rows[row-1][cell-1]

for row in range(blank_start + blank_length, max_row+blank_length+1):
    for cell in range(1, max_col+1):
        sheet.cell(row=row, column=cell).value = rows[row-blank_length-1][cell-1]
        
wb.save('blanked-'+name)

print("A copy of the spreadsheet with blanks inserted has been saved as 'blanked-spreadsheet_name.' It can be found in the same directory as the original.")



