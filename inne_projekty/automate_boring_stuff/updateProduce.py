# updateProduc.py - Corrects costs in produce sales spreadsheet by using
# dictionary with product's corrected prices

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

# The produce types and their updated prices - only here make changes in case of
# new price's updates
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# looping through the rows to update prices, which automatically show new totals
for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1)
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')

